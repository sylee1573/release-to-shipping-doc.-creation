import io
import re

import openpyxl


def extract_text_from_excel(file_bytes: bytes, filename: str = "") -> str:
    """xlsx/xls 파일을 읽어 AI 파싱용 텍스트로 변환."""
    if filename.lower().endswith(".xls"):
        return _extract_xls(file_bytes)
    return _extract_xlsx(file_bytes)


def _extract_xlsx(file_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            if not any(cell is not None for cell in row):
                continue
            cells = [str(cell).strip() if cell is not None else "" for cell in row]
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                rows_text.append(" | ".join(cells))
        if rows_text:
            sections.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_text))
    return "\n\n".join(sections)


# ─── 품목마스터 일괄 업로드 파서 ───────────────────────────────

# 정규화된 헤더 → 필드명. 헤더는 소문자화 + 공백·괄호·/·_ 제거 후 비교.
_ITEM_HEADER_ALIASES = {
    "customer_name":       ["고객사명", "고객사", "발주처", "customer", "customername"],
    "part_number":         ["품번", "품번pn", "부품번호", "pn", "partnumber", "partno", "part"],
    "description":         ["품목설명", "설명", "품명", "품목명", "description", "desc"],
    "unit_price":          ["단가", "단가usd", "price", "unitprice"],
    "net_weight_per_pc":   ["개당순중량", "순중량", "개당중량", "netweight", "netweightperpc"],
    "gross_weight_per_pc": ["개당gross중량", "gross중량", "총중량", "grossweight", "grossweightperpc"],
    "pcs_per_box":         ["박스당수량", "박스당", "박스당pcs", "pcsperbox", "qtyperbox"],
    "boxes_per_pallet":    ["파레트당박스수", "파레트당박스", "파레트당", "boxesperpallet"],
    "cbm_per_pallet":      ["파레트당cbm", "cbm", "cbmperpallet"],
}
_INT_FIELDS = {"pcs_per_box", "boxes_per_pallet"}
_FLOAT_FIELDS = {"unit_price", "net_weight_per_pc", "gross_weight_per_pc", "cbm_per_pallet"}


def _norm_header(value) -> str:
    """헤더 셀 정규화: 소문자화 + 영문/숫자/한글 외 모두 제거 (공백·괄호·단위기호 등)."""
    if value is None:
        return ""
    return re.sub(r"[^0-9a-z가-힣]", "", str(value).strip().lower())


def _match_field(norm_header: str) -> str | None:
    """정규화된 헤더에 포함된 별칭 중 가장 긴 것의 필드 반환 (단위접미사 '개당순중량kg' 등 흡수)."""
    best_field, best_len = None, 0
    for field, aliases in _ITEM_HEADER_ALIASES.items():
        for alias in aliases:
            if alias in norm_header and len(alias) > best_len:
                best_field, best_len = field, len(alias)
    return best_field


def _read_rows(file_bytes: bytes, filename: str) -> list[list]:
    """첫 시트를 행 배열(list[list])로 반환."""
    if filename.lower().endswith(".xls"):
        import xlrd  # xlrd 2.x: .xls 전용
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        return [list(ws.row_values(i)) for i in range(ws.nrows)]
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def parse_item_master_rows(file_bytes: bytes, filename: str) -> tuple[list[dict], str | None]:
    """
    Excel을 품목마스터 dict 리스트로 파싱 (양식 유연 — 헤더 자동탐지·열 순서 무관).
    반환: (rows, error). error가 있으면 rows는 빈 리스트.
    각 row dict: {row: 엑셀행번호, customer_name, part_number, description, ...}
    """
    try:
        raw_rows = _read_rows(file_bytes, filename)
    except Exception as e:  # 손상 파일 등
        return [], f"파일을 읽을 수 없습니다: {e}"

    # 헤더 행 자동탐지 (앞쪽 10행 중 customer_name·part_number 별칭이 둘 다 있는 행)
    header_idx = None
    col_map: dict[int, str] = {}
    for idx in range(min(10, len(raw_rows))):
        candidate: dict[int, str] = {}
        for col, cell in enumerate(raw_rows[idx]):
            norm = _norm_header(cell)
            if not norm:
                continue
            field = _match_field(norm)
            if field and field not in candidate.values():
                candidate[col] = field
        if "customer_name" in candidate.values() and "part_number" in candidate.values():
            header_idx, col_map = idx, candidate
            break

    if header_idx is None:
        return [], "헤더 행을 찾을 수 없습니다. '고객사명'·'품번' 열이 포함된 행이 필요합니다."

    rows: list[dict] = []
    for idx in range(header_idx + 1, len(raw_rows)):
        raw = raw_rows[idx]
        if not any(c is not None and str(c).strip() != "" for c in raw):
            continue  # 빈 행
        item: dict = {"row": idx + 1}  # 1-based 엑셀 행번호
        for col, field in col_map.items():
            val = raw[col] if col < len(raw) else None
            if val is None or str(val).strip() == "":
                continue
            if field in _INT_FIELDS:
                try:
                    item[field] = int(float(str(val).replace(",", "")))
                except (ValueError, TypeError):
                    pass
            elif field in _FLOAT_FIELDS:
                try:
                    item[field] = float(str(val).replace(",", ""))
                except (ValueError, TypeError):
                    pass
            else:
                item[field] = str(val).strip()
        rows.append(item)

    return rows, None


def _extract_xls(file_bytes: bytes) -> str:
    import xlrd  # xlrd 2.x: .xls 전용 (xlrd==2.0.1 required)
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sections = []
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows_text = []
        for row_idx in range(ws.nrows):
            row = ws.row_values(row_idx)
            if not any(cell is not None and cell != "" for cell in row):
                continue
            cells = [str(cell).strip() if cell != "" and cell is not None else "" for cell in row]
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                rows_text.append(" | ".join(cells))
        if rows_text:
            sections.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_text))
    return "\n\n".join(sections)
