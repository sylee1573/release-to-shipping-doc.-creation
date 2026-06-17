import io
import logging
import math
import re
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

# 배포 컨텍스트(backend/) 안에 둬야 Docker(COPY backend/ .)·Railway 배포에 포함됨.
# 루트 inv.packing_ex/ 는 git 미추적이라 배포본에 없었음 → 다운로드 500의 원인이었음.
TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "KCR26-06.xlsx"

# 파레트 1개당 CBM(부피, m³) 표준값 — Packing List CBM = 파레트수 × 이 값
CBM_PER_PALLET = 1.21

SENDER_NAME = "KYUNG CHANG PRECISION IND. CO., LTD."
SENDER_ADDRESS = "149 Gukgasandan-daero 33-gil, Guji-myeon, Dalseong-gun, Daegu"
SENDER_TEL = "82-53-589-0133"
SENDER_FAX = "82-53-582-1786"


# ─── Style helpers (생산의뢰서용) ─────────────────────────────

def _side(style="thin"):
    return Side(style=style)

def _border(left="thin", right="thin", top="thin", bottom="thin"):
    return Border(left=_side(left), right=_side(right), top=_side(top), bottom=_side(bottom))

def _thin():
    return _border()

def _set(ws, row, col, value, bold=False, size=10, align="left", valign="center",
         border=None, fill_color=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=size, name="Arial")
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if border:
        cell.border = border
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    return cell

def _col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ─── 공통 헬퍼 ────────────────────────────────────────────────

def _address_lines(ship_to: str, delivery_location: str, max_lines: int = 8) -> list[str]:
    lines = []
    if ship_to:
        lines.append(ship_to)
    if delivery_location:
        for part in delivery_location.replace("\n", ",").split(","):
            part = part.strip()
            if part:
                lines.append(part)
    return lines[:max_lines]


def _calc_item_amounts(item: dict):
    """(qty_num, up_num, extended) 반환. 단가 없으면 빈 문자열."""
    qty = item.get("quantity", 0)
    unit_price = item.get("unit_price")
    try:
        qty_num = int(qty)
    except (ValueError, TypeError):
        qty_num = qty

    if unit_price is None:
        return qty_num, "", ""
    try:
        up_num = float(unit_price)
        extended = round(qty_num * up_num, 2) if isinstance(qty_num, int) else ""
    except (ValueError, TypeError):
        return qty_num, "", ""
    return qty_num, up_num, extended


def _keep_only(wb, sheet_name: str):
    for sname in list(wb.sheetnames):
        if sname != sheet_name:
            wb.remove(wb[sname])


def _apply_total_style(ws, row: int, n_cols: int):
    """TOTAL 행: 상단 구분선 + 수직 컬럼선, 흰 배경, bold."""
    no   = Side(style=None)
    thin = Side(style="thin")
    med  = Side(style="medium")
    white = PatternFill("solid", fgColor="FFFFFF")

    for c in range(1, max(n_cols, 20) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = white
        cell.border = Border(left=no, right=no, top=no, bottom=no)
        cell.font   = Font(name="Arial", size=11, bold=False)
        if c in (8, 9, 10, 11, 12, 13):
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif c in (14, 15):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18


def _apply_clean_data_rows(ws, start_row: int, count: int, n_cols: int):
    """데이터 행: 모든 라인 없음, 흰 배경, bold. 템플릿 원본 border 포함 제거."""
    no_side   = Side(style=None)
    no_border = Border(left=no_side, right=no_side, top=no_side, bottom=no_side)
    white     = PatternFill("solid", fgColor="FFFFFF")
    clear_cols = max(n_cols, 20)  # 템플릿 잔여 border 포함 충분히 제거

    for i in range(count):
        row = start_row + i
        ws.row_dimensions[row].height = 25
        for col in range(1, clear_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill   = white
            cell.border = no_border
            cell.font   = Font(name="Arial", size=11, bold=False)

def _ran_digits(value) -> int | str:
    """RAN# 값에서 숫자만 추출 (예: 'RAN-001' → 1)."""
    if not value:
        return ""
    m = re.search(r"\d+", str(value))
    return int(m.group()) if m else value


def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header(ws, sheet_type: str, header: dict):
    """
    KCR26-06.xlsx 기준 헤더 셀 채우기.
    IN / PA 시트 셀 위치:
      H2  (col 8,  row 2)  = Invoice/Doc No.
      K2  (col 11, row 2)  = 발행일  (병합 K2:M2)
      A4~A9                = Consignee 주소 (6줄)
      I16 / H16            = SA# (IN: col 9, PA: col 8)
      D18 (col 4,  row 18) = Final Destination  (병합 D18:F18)
      I18 / H18            = Tax ID (IN: col 9, PA: col 8)
      D23 (col 4,  row 23) = Sailing Date  (병합 D23:F23)
    """
    # 서류 번호·발행일
    ws.cell(row=2, column=8).value  = header.get("doc_number", "")
    ws.cell(row=2, column=11).value = date.today()

    # 수신처 주소 (rows 4~9, 최대 6줄)
    addr = _address_lines(
        header.get("ship_to_name", header.get("customer_name", "")),
        header.get("delivery_location", ""),
        max_lines=6,
    )
    for i, row_num in enumerate(range(4, 10)):
        ws.cell(row=row_num, column=1).value = addr[i] if i < len(addr) else ""

    # SA# — IN: I16 (col=9), PA: H16 (col=8)
    sa_col = 9 if sheet_type == "invoice" else 8
    ws.cell(row=16, column=sa_col).value = f"  SA# : {header.get('po_number', '')}"

    # Final Destination — D18 (병합 D18:F18), 공통
    ws.cell(row=18, column=4).value = header.get("final_destination", "")

    # Tax ID — 빈 값으로 초기화 (IN: I18, PA: H18)
    tax_col = 9 if sheet_type == "invoice" else 8
    ws.cell(row=18, column=tax_col).value = None

    # Sailing Date — D23 (병합 D23:F23), 공통
    ws.cell(row=23, column=4).value = header.get("sailing_date", "")


# ─── 생산의뢰서 (내부 문서 — 코드 기반) ─────────────────────

def build_production_request(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "생산의뢰서"

    # 제목
    ws.merge_cells("A1:H1")
    _set(ws, 1, 1, "생산의뢰서 (4주 롤링 계획)", bold=True, size=16, align="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:H2")
    _set(ws, 2, 1, SENDER_NAME, size=10, align="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    # 기본 정보
    info_rows = [
        ("의뢰서 번호",      data.get("request_number", "")),
        ("발주처 (고객사)",   data.get("customer_name", "")),
        ("품번 (P/N)",       data.get("part_number", "")),
        ("품목 설명",        data.get("description", "")),
        ("발주번호 (P.O.#)", data.get("po_number", "")),
        ("내부 RAN#",        str(data.get("ran_number", ""))),
    ]
    for i, (label, value) in enumerate(info_rows, start=4):
        ws.row_dimensions[i].height = 20
        _set(ws, i, 1, label, bold=True, size=10, border=_thin(), fill_color="E8F4FD")
        ws.merge_cells(f"B{i}:H{i}")
        _set(ws, i, 2, value, size=10, border=_thin())

    # 4주 롤링 스케줄 테이블
    sched_start = 4 + len(info_rows) + 1
    ws.row_dimensions[sched_start - 1].height = 6

    # 헤더
    hdr = sched_start
    ws.row_dimensions[hdr].height = 22
    hdr_fill = "4472C4"
    for col, label in enumerate(["슬롯", "주간 시작일", "납품일", "생산수량(EA)", "선적 예정일", "생산완료일", "비고"], start=1):
        cell = ws.cell(row=hdr, column=col, value=label)
        cell.font = Font(bold=True, size=9, name="Arial", color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin()
        cell.fill = PatternFill("solid", fgColor=hdr_fill)

    weekly = data.get("weekly_schedule") or []
    for si, slot in enumerate(weekly):
        r = hdr + 1 + si
        ws.row_dimensions[r].height = 18
        is_hol = slot.get("is_holiday", False)
        row_fill = "FFF2CC" if is_hol else None
        vals = [
            f"{slot.get('slot', si+1)}주차",
            slot.get("week_start", ""),
            slot.get("delivery_date", ""),
            "" if is_hol else slot.get("quantity", ""),
            "" if is_hol else slot.get("sailing_date", ""),
            "" if is_hol else slot.get("production_end", ""),
            slot.get("holiday_reason", "") if is_hol else "",
        ]
        for col, v in enumerate(vals, start=1):
            _set(ws, r, col, v, size=9, border=_thin(),
                 fill_color=row_fill, align="center" if col in (1, 4) else "left")

    # 변경이력
    history = data.get("change_history", [])
    if history:
        base = hdr + 1 + max(len(weekly), 4) + 2
        ws.merge_cells(f"A{base}:H{base}")
        _set(ws, base, 1, "변경 이력", bold=True, size=10, border=_thin(), fill_color="FFF2CC")
        for j, h in enumerate(history, start=1):
            r2 = base + j
            ws.merge_cells(f"A{r2}:H{r2}")
            txt = (
                f"[{h.get('changed_at','')[:10]}] {h.get('field','')}:"
                f" {h.get('before','')} → {h.get('after','')}  ({h.get('reason','')})"
            )
            _set(ws, r2, 1, txt, size=9, border=_thin())
            ws.row_dimensions[r2].height = 16

    _col_widths(ws, {"A": 10, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 20})
    return _to_bytes(wb)


# ─── Commercial Invoice (다품번 지원) ────────────────────────

def build_invoice(header: dict, items: list[dict] | None = None) -> bytes:
    """
    header: doc_number, ship_to_name, delivery_location, final_destination,
            sailing_date, po_number, [tax_id, customer_name]
    items:  list of {part_number, description, quantity, unit_price, po_number, ran_number}
    """
    # 단일 호출 호환 (header에 items 필드 포함된 경우)
    if items is None:
        items = header.get("_items") or [header]

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Invoice 양식 파일 없음: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["IN"]

    _write_header(ws, "invoice", header)

    # 데이터 행 (row 29부터)
    DATA_START = 29
    total_qty = 0
    total_ext = 0.0
    has_price = False

    for idx, item in enumerate(items):
        row = DATA_START + idx
        qty_num, up_num, extended = _calc_item_amounts(item)
        ws[f"A{row}"] = item.get("part_number", "")
        ws[f"E{row}"] = item.get("description", "")
        ws[f"H{row}"] = qty_num
        ws[f"J{row}"] = up_num if up_num != "" else None
        ws[f"K{row}"] = extended if extended != "" else None
        ws[f"M{row}"] = item.get("po_number", header.get("po_number", ""))
        ws[f"N{row}"] = _ran_digits(item.get("ran_number", ""))

        if isinstance(qty_num, int):
            total_qty += qty_num
        if extended != "":
            total_ext += float(extended)
            has_price = True

    # TOTAL 행
    total_row = DATA_START + len(items)
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"H{total_row}"] = total_qty
    ws[f"K{total_row}"] = round(total_ext, 2) if has_price else None
    _apply_clean_data_rows(ws, DATA_START, len(items), 14)
    _apply_total_style(ws, total_row, 14)

    _keep_only(wb, "IN")
    return _to_bytes(wb)


# ─── Packing List (다품번 지원) ──────────────────────────────

def build_packing_list(header: dict, items: list[dict] | None = None) -> bytes:
    """
    header: doc_number, ship_to_name, delivery_location, final_destination,
            sailing_date, po_number
    items:  list of {part_number, description, quantity, net_weight_per_pc,
                     pcs_per_box, po_number, ran_number}
    """
    if items is None:
        items = header.get("_items") or [header]

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Packing List 양식 파일 없음: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["PA"]

    _write_header(ws, "packing", header)

    DATA_START = 29
    total_qty   = 0
    total_box   = 0
    total_net   = 0.0
    total_gross = 0.0
    total_cbm   = 0.0
    has_weight  = False
    has_cbm     = False

    for idx, item in enumerate(items):
        row = DATA_START + idx
        qty            = item.get("quantity", 0)
        pcs_per_box    = item.get("pcs_per_box") or 360
        boxes_per_plt  = item.get("boxes_per_pallet")
        net_w          = item.get("net_weight_per_pc") or 0
        gross_w        = item.get("gross_weight_per_pc")
        cbm_per_pallet = item.get("cbm_per_pallet")

        try:
            qty_num   = int(qty)
            box_count = math.ceil(qty_num / int(pcs_per_box))
            plt_count = math.ceil(box_count / int(boxes_per_plt)) if boxes_per_plt else ""
            net_total = round(float(net_w) * qty_num, 3) if net_w else ""
            if gross_w:
                gross_total = round(float(gross_w) * qty_num, 3)
            elif net_total != "":
                gross_total = round(float(net_total) * 1.023, 3)
            else:
                gross_total = ""
            # CBM = 파레트수 × 파레트당 CBM(품목에 지정값 있으면 우선, 없으면 표준 1.21)
            cbm_rate  = float(cbm_per_pallet) if cbm_per_pallet else CBM_PER_PALLET
            cbm_total = round(cbm_rate * int(plt_count), 4) if plt_count != "" else ""
        except (ValueError, TypeError):
            qty_num, box_count, net_total, gross_total, cbm_total = qty, "", "", "", ""

        ws[f"A{row}"] = item.get("part_number", "")
        ws[f"E{row}"] = item.get("description", "")
        ws[f"H{row}"] = qty_num
        ws[f"J{row}"] = box_count
        ws[f"K{row}"] = net_total   if net_total   != "" else None
        ws[f"L{row}"] = gross_total if gross_total != "" else None
        ws[f"M{row}"] = cbm_total   if cbm_total   != "" else None
        ws[f"N{row}"] = item.get("po_number", header.get("po_number", ""))
        ws[f"O{row}"] = _ran_digits(item.get("ran_number", ""))

        if isinstance(qty_num, int):
            total_qty += qty_num
        if isinstance(box_count, int):
            total_box += box_count
        if net_total != "":
            total_net   += float(net_total)
            total_gross += float(gross_total)
            has_weight = True
        if cbm_total != "":
            total_cbm += float(cbm_total)
            has_cbm = True

    # TOTAL 행
    total_row = DATA_START + len(items)
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"H{total_row}"] = total_qty
    ws[f"J{total_row}"] = total_box
    ws[f"K{total_row}"] = round(total_net,   3) if has_weight else None
    ws[f"L{total_row}"] = round(total_gross, 3) if has_weight else None
    ws[f"M{total_row}"] = round(total_cbm,   4) if has_cbm   else None
    _apply_clean_data_rows(ws, DATA_START, len(items), 15)
    _apply_total_style(ws, total_row, 15)

    _keep_only(wb, "PA")
    return _to_bytes(wb)
