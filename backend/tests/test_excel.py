"""Excel 생성 테스트 — 생산의뢰서·Invoice·Packing List 양식 검증."""
import io
from datetime import date

import openpyxl
import pytest

from services.excel_builder import build_invoice, build_packing_list, build_production_request

_SAMPLE_DATA = {
    "request_number": "PR-202606-0001",
    "customer_name": "현대자동차",
    "part_number": "85310-AA000",
    "quantity": 500,
    "unit": "EA",
    "production_start_date": date(2026, 6, 21),
    "production_end_date": date(2026, 6, 28),
    "delivery_date": date(2026, 6, 30),
    "delivery_location": "울산 1공장",
    "doc_number": "INV-202606-0001",
}


def _load_wb(excel_bytes: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(excel_bytes))


def test_production_request_excel():
    result = build_production_request(_SAMPLE_DATA)
    wb = _load_wb(result)
    ws = wb.active
    assert ws["A1"].value == "생산의뢰서"
    assert any("PR-202606-0001" in str(ws.cell(r, 2).value or "") for r in range(1, 15))


def test_invoice_excel():
    result = build_invoice(_SAMPLE_DATA)
    wb = _load_wb(result)
    ws = wb.active
    assert ws["A1"].value == "INVOICE"
    assert any("INV-202606-0001" in str(ws.cell(r, 2).value or "") for r in range(1, 15))


def test_packing_list_excel():
    result = build_packing_list(_SAMPLE_DATA)
    wb = _load_wb(result)
    ws = wb.active
    assert ws["A1"].value == "PACKING LIST"
    assert ws.cell(3, 1).value == "No."
